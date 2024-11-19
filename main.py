import openpyxl
from openpyxl.styles import Font 
import os
import json
file_list = [s for s in os.listdir(os.curdir) if "xlsx" in s]

def find_common_substring(s1, s2, min_length=3):
    # 建立動態規劃表
    dp = [[0] * (len(s2) + 1) for _ in range(len(s1) + 1)]
    longest = 0  # 最長字串的長度
    end_pos = 0  # 最長字串在 s1 的結尾位置

    # 動態規劃計算
    for i in range(1, len(s1) + 1):
        for j in range(1, len(s2) + 1):
            if s1[i - 1] == s2[j - 1]:
                dp[i][j] = dp[i - 1][j - 1] + 1
                if dp[i][j] > longest:  # 如果找到更長的子字串
                    longest = dp[i][j]
                    end_pos = i

    # 確保字串符合最小長度要求
    if longest >= min_length:
        return s1[end_pos - longest:end_pos]
    else:
        return None  # 沒有符合條件的子字串

def file2dict(sheet_name):
    def get_actual_max_row(sheet_name):
        for row in range(sheet_name.max_row, 0, -1):
            if any(cell.value is not None for cell in sheet_name[row]):
                return row
        return 0  # 若完全空白，回傳 0
    
    # 取得第一列的標題
    headers = [cell.value for cell in sheet_name[1]]
    data_dict = {header: [] for header in headers}
    if "領用人" not in headers and "最後連線" not in headers:
        return {}
    # 檢查哪些儲存格是合併儲存格
    merged_cells = sheet_name.merged_cells
    for row in sheet_name.iter_rows(min_row=2, max_row=get_actual_max_row(sheet_name), values_only=False):
        for header, cell in zip(headers, row):
            value = cell.value
            # 處理合併儲存格，檢查是否為 None
            if value is None:
                for merged_range in merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # 找到合併範圍的第一個儲存格，取得其值
                        value = sheet_name[merged_range.coord.split(":")[0]].value
                        break
            # 如果仍然是 None，填充空字串
            data_dict[header].append(value if value is not None else "")
    return data_dict

def write_json(name, dic):
    with open(name, "w", encoding="utf-8") as json_file:
        json.dump(dic, json_file, indent=4, ensure_ascii=False)
    print("資料已成功寫入", name)

print("紀錄上線時間的檔案")
for i in range(len(file_list)):
    print(i+1,".", file_list[i])
print("請選擇....", end=" ")
time_select = int(input())-1

print("----------------------------------------------")
print("紀錄機器領用人的檔案")
for i in range(len(file_list)):
    print(i+1,".", file_list[i])
print("請選擇....", end=" ")
name_select = int(input())-1

time_file = openpyxl.load_workbook(file_list[time_select])
name_file = openpyxl.load_workbook(file_list[name_select])


# 迭代所有工作表並合併內容
time_sheet = time_file.worksheets
time_dict = file2dict(time_sheet[0])
for i in range(1, len(time_sheet)):
    current_dict = file2dict(time_sheet[i])
    for key in current_dict:
        if key in time_dict:
            time_dict[key].extend(current_dict[key])

name_sheet = name_file.worksheets
name_dict = file2dict(name_sheet[0])

for i in range(1, len(name_sheet)):
    current_dict = file2dict(name_sheet[i])
    for key in current_dict:
        if key in name_dict:
            name_dict[key].extend(current_dict[key])

write_json("output.json", name_dict)

print("----------------------------------------------")
print("請輸入年份及月份 (如 2024 01)")
year, month = map(int, (input().split()))
useless = []
for i in range(len(time_dict["最後連線"])):
    if int(time_dict["最後連線"][i].split("-")[1])+int(time_dict["最後連線"][i].split("-")[0])*12<month+year*12:
        useless.append(time_dict["serialNumber"][i])
print("尚有", len(useless) ,"台未上線")

dic = {}
for i in range(len(useless)):
    try:
        index = name_dict["Serial Number"].index(useless[i])
        key = name_dict["領用人"][index].replace("\n", "") + name_dict["備註"][index].replace("\n", "")
        if key not in dic:
            dic[key] = {"TabletID": []}
        dic[key]["TabletID"].append(name_dict["平板編號"][index])
        # print(useless[i], name_dict["領用人"][index].replace("\n", ""))
    except ValueError:
        print(useless[i], "NONE")



import csv
# 開啟 CSV 檔案
with open('contacts.csv', newline='',encoding="utf-8") as csvfile:
    # 讀取 CSV 檔案內容
    rows = list(csv.reader(csvfile))
    for n in dic:
        ret_row = []
        max_r = ""
        for row in rows:
            row = [row[0], row[-1]]
            result = find_common_substring(row[0], n)
            if result:
                # print((result), (max_r), len(result), len(max_r))
                if len(result)>len(max_r):
                    max_r=result
                    ret_row = row
        if ret_row:
            dic[n]["email"] = ret_row[1]


write_json("output_mail.json", dic)